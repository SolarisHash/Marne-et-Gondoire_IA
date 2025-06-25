# ============================================================================
# SAUVEGARDE EXCEL AVEC FORMATAGE
# mg-platform/mcp_server/tools/ai_agent/output/excel_writer.py
# ============================================================================

"""
Module de sauvegarde Excel avec format SIRET corrigé et colorisation
Responsabilités:
- Format SIRET forcé en texte (zéros de tête)
- Colonnes métadonnées IA ajoutées
- Colorisation rouge pour données IA
- Path : data/processed/AI_ENRICHED_Sample_{session_id}.xlsx
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
from ..core.exceptions import OutputError


class ExcelWriter:
    """Gestionnaire de sauvegarde Excel avec formatage avancé"""
    
    def __init__(self, config: Dict[str, Any], session_id: str):
        self.config = config
        self.session_id = session_id
        self.output_dir = Path(config.get("processed_data_dir", "data/processed"))
        
        # Styles de colorisation
        self.ai_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Rouge clair
        self.ai_font = Font(color="CC0000", bold=True)  # Rouge foncé et gras
        self.meta_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Bleu clair
        self.meta_font = Font(color="0066CC", italic=True)  # Bleu et italique
    
    def save_enriched_results(self, sample_df: pd.DataFrame, enrichment_results: Dict, performance_metrics: Dict) -> str:
        """
        Sauvegarde les résultats enrichis avec format SIRET corrigé
        
        Args:
            sample_df: DataFrame échantillon original
            enrichment_results: Résultats d'enrichissement
            performance_metrics: Métriques de performance
            
        Returns:
            Chemin du fichier sauvegardé
        """
        
        try:
            # Créer le répertoire de sortie
            self.output_dir.mkdir(parents=True, exist_ok=True)
            
            # Préparer le DataFrame enrichi
            enriched_df = self._prepare_enriched_dataframe(sample_df, enrichment_results)
            
            # Nom du fichier de sortie
            output_filename = f"AI_ENRICHED_Sample_{self.session_id}.xlsx"
            output_path = self.output_dir / output_filename
            
            # Sauvegarder avec formatage SIRET
            self._save_with_siret_formatting(enriched_df, output_path)
            
            # Colorisation si activée
            if self.config.get("excel_colorization", True):
                colorized_path = self._apply_colorization(
                    output_path, enrichment_results["enrichment_data"]
                )
                return str(colorized_path)
            
            return str(output_path)
            
        except Exception as e:
            raise OutputError(f"Erreur sauvegarde Excel: {str(e)}")
    
    def _prepare_enriched_dataframe(self, sample_df: pd.DataFrame, enrichment_results: Dict) -> pd.DataFrame:
        """Prépare le DataFrame avec données enrichies et métadonnées"""
        
        enriched_df = sample_df.copy()
        
        # Correction format SIRET
        self._fix_siret_format(enriched_df)
        
        # Ajouter colonnes de métadonnées IA
        enriched_df["IA_Enriched"] = False
        enriched_df["IA_Confidence_Score"] = 0.0
        enriched_df["IA_Source"] = ""
        enriched_df["IA_Processing_Date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        enriched_df["IA_Session_ID"] = self.session_id
        
        # Appliquer les enrichissements
        for idx_str, enrichment_data in enrichment_results["enrichment_data"].items():
            idx = int(idx_str) - 1  # Convertir index (1-based vers 0-based)
            
            if idx < len(enriched_df):
                # Enrichir site web si disponible
                if "website" in enrichment_data:
                    if "Site Web établissement" in enriched_df.columns:
                        enriched_df.iloc[idx, enriched_df.columns.get_loc("Site Web établissement")] = enrichment_data["website"]
                
                # Métadonnées IA
                enriched_df.iloc[idx, enriched_df.columns.get_loc("IA_Enriched")] = True
                
                # Score de confiance
                if idx_str in enrichment_results["quality_reports"]:
                    score = enrichment_results["quality_reports"][idx_str]["quality_score"]
                    enriched_df.iloc[idx, enriched_df.columns.get_loc("IA_Confidence_Score")] = score
                
                # Source
                source = enrichment_data.get("search_source", "AI_Generated")
                enriched_df.iloc[idx, enriched_df.columns.get_loc("IA_Source")] = source
        
        return enriched_df
    
    def _fix_siret_format(self, df: pd.DataFrame):
        """Corrige le format SIRET pour préserver les zéros de tête"""
        
        # Identifier les colonnes SIRET/SIREN
        siret_columns = []
        for col in df.columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['siret', 'siren']):
                siret_columns.append(col)
        
        # Convertir en texte avec formatage
        for col in siret_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: 
                    x.zfill(14) if x.isdigit() and len(x) <= 14 else x
                )
    
    def _save_with_siret_formatting(self, df: pd.DataFrame, output_path: Path):
        """Sauvegarde avec format SIRET forcé en texte"""
        
        # Sauvegarder avec ExcelWriter pour contrôler le formatage
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Données Enrichies')
            
            # Récupérer le workbook et worksheet
            workbook = writer.book
            worksheet = writer.sheets['Données Enrichies']
            
            # Créer style texte
            text_style = NamedStyle(name="text_format")
            text_style.number_format = '@'  # Format texte
            
            # Appliquer le format aux colonnes SIRET
            siret_columns = [col for col in df.columns if 'siret' in col.lower() or 'siren' in col.lower()]
            
            for col_name in siret_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # Excel commence à 1
                    
                    # Appliquer à toute la colonne
                    for row in range(1, len(df) + 2):  # +2 pour header et index
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.style = text_style
                        
                        # Pour les cellules de données, s'assurer que c'est bien du texte
                        if row > 1:  # Pas le header
                            cell_value = str(cell.value)
                            if cell_value.isdigit():
                                cell.value = cell_value.zfill(14)  # Pad avec zéros
    
    def _apply_colorization(self, file_path: Path, enrichment_data: Dict) -> Path:
        """Applique la colorisation pour différencier les données IA"""
        
        try:
            # Charger le workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Identifier les colonnes enrichies
            enriched_columns = self._find_enriched_columns(ws)
            
            # Coloriser les cellules enrichies
            for row_idx, enrichment in enrichment_data.items():
                excel_row = int(row_idx) + 1  # +1 pour header, index commence à 1
                
                # Coloriser les données enrichies
                self._colorize_enriched_row(ws, excel_row, enriched_columns, enrichment)
            
            # Coloriser les colonnes de métadonnées IA
            self._colorize_metadata_columns(ws)
            
            # Ajouter légende
            self._add_legend(ws)
            
            # Sauvegarder le fichier colorisé
            colorized_path = file_path.with_name(file_path.stem + '_COLORIZED.xlsx')
            wb.save(colorized_path)
            
            return colorized_path
            
        except Exception as e:
            # Retourner le fichier original si colorisation échoue
            print(f"⚠️ Erreur colorisation: {e}")
            return file_path
    
    def _find_enriched_columns(self, ws) -> list:
        """Trouve les colonnes qui ont été enrichies"""
        enriched_cols = []
        
        # Parcourir la première ligne (header)
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_name = str(cell.value).lower()
                
                # Colonnes typiquement enrichies
                if any(keyword in col_name for keyword in [
                    'site', 'web', 'email', 'telephone', 'phone',
                    'linkedin', 'facebook', 'twitter', 'url'
                ]):
                    enriched_cols.append(col_idx)
        
        return enriched_cols
    
    def _colorize_enriched_row(self, ws, row_idx: int, enriched_columns: list, enrichment_data: dict):
        """Colorise une ligne enrichie par IA"""
        
        for col_idx in enriched_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Vérifier si la cellule contient des données IA
            if self._cell_contains_ai_data(cell, enrichment_data):
                # Appliquer le style IA (rouge)
                cell.fill = self.ai_fill
                cell.font = self.ai_font
    
    def _cell_contains_ai_data(self, cell, enrichment_data: dict) -> bool:
        """Vérifie si une cellule contient des données enrichies par IA"""
        
        if not cell.value or not enrichment_data:
            return False
        
        cell_value = str(cell.value).strip()
        
        # Vérifier si la valeur correspond aux données IA
        ai_values = []
        if 'website' in enrichment_data:
            ai_values.append(enrichment_data['website'])
        if 'company_name' in enrichment_data:
            ai_values.append(enrichment_data['company_name'])
        
        return any(cell_value == str(ai_val) for ai_val in ai_values if ai_val)
    
    def _colorize_metadata_columns(self, ws):
        """Colorise les colonnes de métadonnées IA"""
        
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and str(cell.value).startswith('IA_'):
                # Coloriser toute la colonne de métadonnées
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.fill = self.meta_fill
                        cell.font = self.meta_font
    
    def _add_legend(self, ws):
        """Ajoute une légende explicative"""
        
        # Trouver la dernière ligne avec données
        last_row = ws.max_row + 2
        
        # Ajouter la légende
        legend_data = [
            ["", "LÉGENDE"],
            ["", "🔴 Rouge = Données enrichies par IA"],
            ["", "🔵 Bleu = Métadonnées IA"],
            ["", "⚪ Standard = Données originales"]
        ]
        
        for i, (col1, col2) in enumerate(legend_data):
            ws.cell(row=last_row + i, column=1, value=col1)
            legend_cell = ws.cell(row=last_row + i, column=2, value=col2)
            
            if i == 0:  # Titre
                legend_cell.font = Font(bold=True, size=12)
            elif "Rouge" in col2:
                legend_cell.fill = self.ai_fill
                legend_cell.font = self.ai_font
            elif "Bleu" in col2:
                legend_cell.fill = self.meta_fill
                legend_cell.font = self.meta_font